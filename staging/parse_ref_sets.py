from collections import OrderedDict
from contextlib import contextmanager
from numbers import Number
from recordclass import recordclass
from unicodecsv import writer as csv_writer

from parse_fts import oracle_ctl_csv, oracle_ddl, load_script


DATE = 'DATE'
NUMBER = 'NUMBER'
VARCHAR2 = 'VARCHAR2'
MIN_VARCHAR2_LEN = 8

ColType = recordclass('ColType', 'typ max_len')


def main(load_workbook_argv, open_wr_cwd, get_cli, datetime,
         sqlldr_script='sqlldr_all_ref.sh',
         sql_create='oracle_create_ref.sql',
         sql_drop='oracle_drop_ref.sql'):
    def err(typ_str, old_typ_str, sheet_name, idx):
        raise RuntimeError('%s column changed to %s! %s, %d' %
                           (old_typ_str, typ_str, sheet_name, idx))

    def p2size(sz):
        return 1 << (sz-1).bit_length()

    def col_xlate(c):
        ''' Replace Oracle reserved words.
        '''
        return 'levl' if c == 'level' else c

    load_script_data = 'set -evx\n\n'
    sql_data = ''
    tables = []

    wb = load_workbook_argv(get_cli()[1])
    for sheet_name in wb.get_sheet_names():
        sh = wb.get_sheet_by_name(sheet_name)
        table_name = 'ref_' + sheet_name.replace(' ', '_').lower()
        tables.append(table_name)

        print 'Processing %s' % table_name

        csv_file_name = table_name + '.csv'
        with open_wr_cwd(csv_file_name) as fout:
            w = csv_writer(fout)
            header = None
            for idx, row in enumerate(sh.rows):
                # The first few lines may be description/title
                if not header:
                    if None not in [c.value for c in row]:
                        header = OrderedDict([(
                            col_xlate(cell.value.replace(' ', '_').lower()),
                            ColType(None, MIN_VARCHAR2_LEN))
                                              for cell in row])
                else:
                    row_to_write = []
                    for cell, head in zip(row, header.keys()):
                        if isinstance(cell.value, datetime):
                            if header[head].typ and header[head].typ != DATE:
                                err(DATE, header[head].typ, sheet_name, idx)
                            header[head].typ = DATE
                            row_to_write.append(cell.value.strftime('%Y%m%d'))
                        elif isinstance(cell.value, Number):
                            if header[head].typ and header[head].typ != NUMBER:
                                err(NUMBER, header[head].typ, sheet_name, idx)
                            header[head].typ = NUMBER
                            row_to_write.append(cell.value)
                        elif cell.value:
                            if(header[head].typ and
                               header[head].typ != VARCHAR2):
                                err(VARCHAR2, header[head].typ,
                                    sheet_name, idx)
                            header[head].typ = VARCHAR2
                            header[head].max_len = (
                                max(header[head].max_len,
                                    p2size(len(cell.value) +
                                           MIN_VARCHAR2_LEN)))
                            row_to_write.append(cell.value.strip()
                                                if cell.value.strip()
                                                else None)
                        else:
                            row_to_write.append(None)
                    if True in [c is not None for c in row_to_write]:
                        w.writerow(row_to_write)

        ctl_file_name = write_ctl(table_name, header, open_wr_cwd)
        sql_data += sql(table_name, header) + '\n\n'
        load_script_data += load_script(ctl_file_name, csv_file_name,
                                        csv_file_name)

    with open_wr_cwd(sqlldr_script) as fout:
        fout.write(load_script_data)
    with open_wr_cwd(sql_create) as fout:
        fout.write(sql_data)
    with open_wr_cwd(sql_drop) as fout:
        fout.write('\n'.join(['drop table %s;' % t for t in tables]))


def write_ctl(table_name, header, open_wr_cwd):
    fn = table_name + '.ctl'
    with open_wr_cwd(fn) as fout:
        cols = list()
        for cname, ct in header.items():
            if ct.typ == DATE:
                cols.append(cname + ' ' + ct.typ + " 'yyyymmdd'")
            elif ct.typ == VARCHAR2:
                cols.append(cname + ' char(%d)' % ct.max_len)
            else:
                cols.append(cname)
        fout.write(oracle_ctl_csv(table_name, cols))
    return fn


def sql(table_name, header):
    return oracle_ddl(table_name, [(cname + ' ' + ct.typ +
                                    ('(%d)' % ct.max_len
                                     if ct.typ == VARCHAR2 else ''))
                                   for (cname, ct) in header.items()])


if __name__ == '__main__':
    def _tcb():
        from datetime import datetime
        from openpyxl import load_workbook
        from os import getcwd
        from os.path import abspath
        from sys import argv

        def get_input_path():
            return argv[1]

        def get_cli():
            return argv

        def load_workbook_argv(path):
            if get_input_path() not in path:
                raise RuntimeError('%s not in %s' % (get_input_path(), path))
            return load_workbook(path, read_only=True)

        @contextmanager
        def open_wr_cwd(path):
            cwd = abspath(getcwd())
            if cwd not in abspath(path):
                raise RuntimeError('%s not in %s' % (cwd, path))
            with open(path, 'wb') as fin:
                yield fin

        main(load_workbook_argv, open_wr_cwd, get_cli, datetime)

    _tcb()
